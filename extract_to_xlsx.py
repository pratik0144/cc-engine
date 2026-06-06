#!/usr/bin/env python3
"""
Extract all cached HDFC cards and write to the xlsx spreadsheet.
Runs LLM extraction on each cached file, then maps fields to xlsx columns.
"""
import asyncio
import json
import os
import sys
from datetime import datetime, timezone

import openpyxl
from engine.extractor import extract_from_cached_file, check_ollama_health
from engine.normalizer import normalize_card
from engine.utils import logger, generate_card_id

XLSX_PATH = "/Users/pratikpotadar/Developer/cc.com/Untitled spreadsheet.xlsx"
RAW_DIR = "data/raw/hdfc"

# Column mapping: xlsx column number → field name from our extraction
COLUMN_MAP = {
    1: "bank_name",
    2: "card_name",
    3: "card_id",
    4: "card_type",
    5: "card_network",
    6: "variant",
    7: "card_category",
    8: "source_url",
    9: "last_verified_at",
    # 10: spacer
    11: "joining_fee",
    12: "annual_fee",
    13: "late_payment_fee",
    14: "cash_withdrawal_fee",
    15: "reward_redemption_fee",
    16: "forex_markup",
    # 17: spacer
    18: "reward_type",
    19: "reward_rate_general",
    20: "reward_rate_category_wise",
    21: "_all_bonus",       # Combined: welcome_bonus + milestone_bonuses
    22: "fuel_surcharge_waiver",
    23: "domestic_lounge_visits",   # mapped from lounge_access or domestic_lounge_visits
    24: "international_lounge_visits",
    25: "_all_offers",      # Combined: movie_offer + dining_offer + travel_benefits
    26: "contactless_support",
    # 27: spacer
    28: "minimum_income",
    29: "employment_type",
    30: "_age_range",       # Combined: age_min & age_max
    31: "city_tier",
    32: "cibil_range_hint",
    33: "salary_account_required",
    34: None,  # card_img — skip
}


def build_row_data(card: dict) -> dict:
    """Build a dict with column numbers as keys from extracted card data."""
    row = {}
    
    for col_num, field_name in COLUMN_MAP.items():
        if field_name is None:
            continue
        
        if field_name == "_all_bonus":
            # Combine welcome_bonus + milestone_bonuses
            parts = []
            if card.get("welcome_bonus"):
                parts.append(f"Welcome: {card['welcome_bonus']}")
            if card.get("milestone_bonuses"):
                mb = card["milestone_bonuses"]
                if isinstance(mb, list):
                    parts.append(f"Milestone: {'; '.join(str(x) for x in mb)}")
                elif mb:
                    parts.append(f"Milestone: {mb}")
            row[col_num] = " | ".join(parts) if parts else None
            
        elif field_name == "_all_offers":
            # Combine movie + dining + travel + insurance + grocery
            parts = []
            for f, label in [
                ("movie_offer", "Movie"), ("dining_offer", "Dining"),
                ("travel_benefits", "Travel"), ("insurance_benefits", "Insurance"),
                ("grocery_offer", "Grocery"),
            ]:
                if card.get(f):
                    parts.append(f"{label}: {card[f]}")
            row[col_num] = " | ".join(parts) if parts else None
            
        elif field_name == "_age_range":
            # Combine age_min & age_max
            age_min = card.get("age_min")
            age_max = card.get("age_max")
            if age_min or age_max:
                row[col_num] = f"{age_min or '?'} - {age_max or '?'}"
            else:
                row[col_num] = None
                
        elif field_name == "reward_rate_category_wise":
            val = card.get(field_name)
            if isinstance(val, dict):
                row[col_num] = "; ".join(f"{k}: {v}" for k, v in val.items())
            elif val:
                row[col_num] = str(val)
            else:
                row[col_num] = None
                
        elif field_name == "domestic_lounge_visits":
            # Try specific field first, then general lounge_access
            val = card.get("domestic_lounge_visits")
            if not val and card.get("lounge_access"):
                val = card["lounge_access"]
            row[col_num] = val
            
        else:
            val = card.get(field_name)
            if isinstance(val, list):
                row[col_num] = "; ".join(str(x) for x in val)
            elif val is not None:
                row[col_num] = str(val)
            else:
                row[col_num] = None
    
    return row


async def main():
    # Check Ollama
    if not await check_ollama_health():
        print("❌ Ollama not available!")
        sys.exit(1)
    
    # Find all cached files
    files = sorted([
        f for f in os.listdir(RAW_DIR)
        if f.endswith('.json') and f not in ('_discovered_urls.json', 'listing-page.json')
    ])
    print(f"\n📂 Found {len(files)} cached HDFC card files")
    
    # Extract all cards
    all_cards = []
    failed = []
    
    for i, filename in enumerate(files, 1):
        path = os.path.join(RAW_DIR, filename)
        print(f"\n[{i}/{len(files)}] {filename}")
        
        try:
            cards = await extract_from_cached_file(path)
            if cards:
                # Normalize
                for card in cards:
                    card = normalize_card(card, "hdfc")
                    all_cards.append(card)
                print(f"  ✅ Extracted: {cards[0].get('card_name', '?')} ({sum(1 for v in cards[0].values() if v is not None)} fields)")
            else:
                failed.append(filename)
                print(f"  ❌ FAILED")
        except Exception as e:
            failed.append(filename)
            print(f"  ❌ ERROR: {e}")
    
    print(f"\n{'='*60}")
    print(f"📊 Results: {len(all_cards)} cards extracted, {len(failed)} failed")
    print(f"{'='*60}")
    
    if not all_cards:
        print("No cards to write!")
        return
    
    # Write to xlsx
    print(f"\n📝 Writing to xlsx: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["Sheet1"]
    
    # Clear existing data rows (keep header)
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None
    
    # Write card data
    for idx, card in enumerate(all_cards):
        row_num = idx + 2  # Row 1 is header
        row_data = build_row_data(card)
        
        for col_num, value in row_data.items():
            if value is not None:
                ws.cell(row=row_num, column=col_num).value = value
    
    # Update Sheet2 bank tracking
    if "Sheet2" in wb.sheetnames:
        ws2 = wb["Sheet2"]
        for row in range(2, ws2.max_row + 1):
            bank_name = ws2.cell(row=row, column=2).value
            if bank_name and "hdfc" in str(bank_name).lower():
                ws2.cell(row=row, column=3).value = len(all_cards)
                ws2.cell(row=row, column=4).value = "✅ Done"
                break
    
    wb.save(XLSX_PATH)
    print(f"✅ Saved {len(all_cards)} cards to xlsx!")
    
    # Print summary
    print(f"\n📋 Cards written:")
    for card in all_cards:
        non_null = sum(1 for v in card.values() if v is not None)
        print(f"  • {card.get('card_name', '?')} ({non_null} fields)")
    
    if failed:
        print(f"\n❌ Failed files ({len(failed)}):")
        for f in failed:
            print(f"  • {f}")


if __name__ == "__main__":
    asyncio.run(main())
