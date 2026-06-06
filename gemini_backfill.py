import os
import time
import json
import openpyxl
import google.generativeai as genai

# Setup API Key provided by user
# Make sure to set the GEMINI_API_KEY environment variable
genai.configure(api_key=os.getenv("GEMINI_API_KEY"))

# Using gemini-flash-latest because 2.5-flash has a strict 20 req/day limit
model = genai.GenerativeModel(
    "gemini-flash-latest",
    generation_config={"response_mime_type": "application/json"}
)

XLSX_PATH = "/Users/pratikpotadar/Developer/cc.com/Untitled spreadsheet.xlsx"

# The user requested to ONLY fill "important" fields to save API quota.
IMPORTANT_COLS = {
    "joining_fee", "annual_fee", "minimum_income", 
    "employment_type", "age_min & age_max", "cibil_range_hint",
    "reward_rate_general"
}

def main():
    print(f"Loading workbook: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["Sheet1"]
    
    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    
    cards_to_process = []
    
    # 1. Collect all rows that need processing
    for row in range(2, ws.max_row + 1):
        bank_name = ws.cell(row=row, column=1).value
        card_name = ws.cell(row=row, column=2).value
        
        if not bank_name or not card_name:
            continue
            
        missing_important_fields = []
        for col, header in enumerate(headers, 1):
            if header in IMPORTANT_COLS:
                val = ws.cell(row=row, column=col).value
                if val is None or str(val).strip() == '' or str(val).lower() == 'none':
                    missing_important_fields.append((col, header))
        
        if missing_important_fields:
            cards_to_process.append({
                "row_idx": row,
                "bank_name": bank_name,
                "card_name": card_name,
                "missing": missing_important_fields
            })

    if not cards_to_process:
        print("No important missing fields found!")
        return
        
    print(f"Found {len(cards_to_process)} cards with missing important fields. Batching calls...")
    
    # 2. Process in batches of 10 to avoid RPM limits completely
    BATCH_SIZE = 10
    
    for i in range(0, len(cards_to_process), BATCH_SIZE):
        batch = cards_to_process[i:i+BATCH_SIZE]
        print(f"\nProcessing batch {i//BATCH_SIZE + 1} ({len(batch)} cards)...")
        
        prompt_lines = [
            "You are an expert on Indian Credit Cards.",
            "I need you to provide the missing data for several credit cards.",
            "If you are absolutely sure, provide the data. If not, provide a reasonable estimate and append ' (est)'. If not applicable, use 'NA'.",
            "Return a JSON object where keys are the Card Names, and the values are objects mapping the missing field names to your answers.",
            "Here are the cards:"
        ]
        
        for card in batch:
            missing_names = [f[1] for f in card['missing']]
            prompt_lines.append(f"\nCard: {card['card_name']} (Bank: {card['bank_name']})")
            prompt_lines.append(f"Missing Fields: {', '.join(missing_names)}")
            
        prompt = "\n".join(prompt_lines)
        
        import time
        max_retries = 5
        for attempt in range(max_retries):
            try:
                response = model.generate_content(prompt)
                data = json.loads(response.text)
                
                updates = 0
                for card in batch:
                    cname = card['card_name']
                    # Gemini might slightly alter the card name key, so try exact then substring
                    matching_key = None
                    for k in data.keys():
                        if cname.lower() in k.lower() or k.lower() in cname.lower():
                            matching_key = k
                            break
                            
                    if matching_key:
                        card_data = data[matching_key]
                        for col, header in card['missing']:
                            if header in card_data:
                                val = str(card_data[header])
                                if val and val != "null" and val.lower() != "none":
                                    ws.cell(row=card['row_idx'], column=col).value = val
                                    updates += 1
                                    
                wb.save(XLSX_PATH)
                print(f"  ✅ Batch successful. Updated {updates} fields.")
                break # Success, exit retry loop
                
            except Exception as e:
                print(f"  ❌ Error processing batch (attempt {attempt+1}/{max_retries}): {e}")
                if "429" in str(e) or "quota" in str(e).lower():
                    wait_time = 30 * (attempt + 1)
                    print(f"  ⏳ Quota exceeded, sleeping for {wait_time}s...")
                    time.sleep(wait_time)
                else:
                    break # Not a rate limit error, don't retry
            
        if i + BATCH_SIZE < len(cards_to_process):
            print("  Waiting 15 seconds to respect rate limits between batches...")
            time.sleep(15)
            
    print("\n🎉 Important backfill complete!")

if __name__ == "__main__":
    main()
